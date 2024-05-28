# tasks.py
from celery import shared_task
from django.core.mail import send_mail
from django.conf import settings
import pandas as pd
from django.conf import settings
from openpyxl import load_workbook

@shared_task
def send_email_task(subject, message, recipient_list):
    try:
        send_mail(
            subject,
            message,
            'rentaluav@gmail.com',
            recipient_list,
            fail_silently=False,
        )
        print("E-posta başarıyla gönderildi.")
    except Exception as e:
        print(f"E-posta gönderimi sırasında bir hata oluştu: {str(e)}")



@shared_task
def save_to_excel(brand_model, username,begin_date,end_date):
    try:
        excel_path = os.path.join(settings.MEDIA_ROOT, 'rentals.xlsx')
        rental_data = {
            'Marka Model': [brand_model],
            'Username': [username],
            'BeginDate': [begin_date],
            'EndDate': [end_date],
        }

        df = pd.DataFrame(rental_data)

        if os.path.exists(excel_path):
            # Mevcut dosyaya ekleme yap
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
                book = load_workbook(excel_path)
                writer.book = book

                # 'Sheet1' sayfasını al
                if 'Sheet1' in book.sheetnames:
                    startrow = book['Sheet1'].max_row
                else:
                    startrow = 0

                df.to_excel(writer, index=False, header=False, startrow=startrow)
        else:
            # Yeni dosya oluştur ve yaz
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
    except Exception as e:
        print(f"Excel dosyasına yazarken hata oluştu: {e}")